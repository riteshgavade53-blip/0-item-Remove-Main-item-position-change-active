from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import io
import os

app = Flask(__name__)

RED_FILL = 'FFEA9999'
ORANGE_FILL = 'FFFFFF00'


# ── HELPERS ─────────────────────────────────────────────

def get_ws_columns(ws):
    """Map worksheet headers to 1-based column indexes."""
    columns = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header is not None:
            columns[str(header).strip()] = col_idx
    return columns


def clean_item_name(name):
    if pd.isna(name):
        return ""
    name = str(name)
    for sym in ['&', '/', '+']:
        name = name.replace(sym, f' {sym} ')
    name = " ".join(name.split())
    name = name.title()
    return name


def safe_price(value):
    """Return a numeric price, falling back to 0 for blanks and NaN."""
    numeric = pd.to_numeric(value, errors='coerce')
    if pd.isna(numeric):
        return 0.0
    return float(numeric)


def get_protected_variation_rows(ws, col_map=None):
    """Rows with orange price cells and a positive variation price must stay unchanged."""
    col_map = col_map or get_ws_columns(ws)
    variation_col = col_map.get('Variation', 4)
    price_col = col_map.get('Price', 8)
    protected = set()
    for row_idx in range(2, ws.max_row + 1):
        variation = ws.cell(row_idx, variation_col).value
        price = ws.cell(row_idx, price_col).value
        fill = ws.cell(row_idx, price_col).fill
        fill_rgb = (
            fill.fgColor.rgb
            if fill and fill.fgColor and fill.fgColor.type == 'rgb'
            else '00000000'
        )
        price_value = safe_price(price)
        if variation and price_value > 0 and fill_rgb == ORANGE_FILL:
            protected.add(row_idx - 2)
    return protected


def get_row_bg(ws, row_idx, preferred_cols=None):
    """Get background color of a row — check col 3 (Item Name) as it's most reliably colored."""
    for col in (preferred_cols or [3, 1, 2]):
        cell = ws.cell(row=row_idx, column=col)
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb != '00000000':
            return fill.fgColor.rgb
    return '00000000'


def set_row_bg(ws, row_idx, color_rgb, num_cols):
    """Set background color for all cells in a row."""
    if color_rgb == '00000000':
        fill = PatternFill(fill_type=None)
    else:
        fill = PatternFill(fill_type='solid', fgColor=color_rgb)
    for c in range(1, num_cols + 1):
        # Don't overwrite the orange price cell (FFFFFF00)
        existing = ws.cell(row_idx, c).fill
        existing_rgb = existing.fgColor.rgb if existing and existing.fgColor and existing.fgColor.type == 'rgb' else '00000000'
        if existing_rgb != 'FFFFFF00':
            ws.cell(row=row_idx, column=c).fill = fill


def get_cell_bg(ws, row_idx, col_idx):
    """Get background color of a single cell."""
    fill = ws.cell(row=row_idx, column=col_idx).fill
    if fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb != '00000000':
        return fill.fgColor.rgb
    return '00000000'


def set_cell_bg(ws, row_idx, col_idx, color_rgb):
    """Set a single cell background color."""
    if color_rgb == '00000000':
        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type=None)
    else:
        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type='solid', fgColor=color_rgb)


def is_orange_price_row(ws, row_idx, col_map=None):
    """Yellow/orange highlighted price rows must remain untouched."""
    col_map = col_map or get_ws_columns(ws)
    price_cell = ws.cell(row_idx, col_map.get('Price', 8))
    fill = price_cell.fill
    if fill and fill.fgColor and fill.fgColor.type == 'rgb':
        return fill.fgColor.rgb == ORANGE_FILL
    return False


# ── MAIN LOGIC ─────────────────────────────────────────────
def process_logic(df, protected_rows=None):
    def normalize_var(v):
        if pd.isna(v):
            return ""
        v = str(v).lower().strip()
        v = v.replace("_", " ").replace("-", " ")
        v = " ".join(v.split())
        return v

    protected_rows = protected_rows or set()

    df['Item Name'] = df['Item Name'].apply(clean_item_name)
    if 'Price' in df.columns:
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0)
    if '__base_name' in df.columns:
        base_group = df['__base_name'].fillna("").astype(str).str.strip()
        df['group'] = base_group.where(base_group.ne(""), df['Item Name'])
    else:
        df['group'] = (df['Item Name'] != df['Item Name'].shift()).cumsum()

    actions = ["KEEP"] * len(df)

    for g in df['group'].unique():
        group_rows = df[df['group'] == g]

        group_rows = group_rows.copy()
        group_rows['_norm_var'] = group_rows['Variation'].apply(normalize_var)
        protected_in_group = [i for i in group_rows.index if i in protected_rows]

        if protected_in_group:
            continue

        veg_rows = group_rows[group_rows['_norm_var'].isin(['veg', 'non veg'])]
        parent_rows = group_rows[group_rows['Price'] == 0]

        if len(veg_rows) > 0 and any(parent_rows['Price'] == 0):
            for i in parent_rows.index:
                if df.loc[i, 'Price'] == 0:
                    actions[i] = "DELETE"
            for i, row in veg_rows.iterrows():
                if row['Price'] > 0:
                    df.at[i, 'Variation'] = ""
                    actions[i] = "KEEP"
            continue

        # Skip if any row has Item ID
        if 'Item ID' in group_rows.columns and group_rows['Item ID'].notna().any():
            continue

        has_variation = group_rows['Variation'].notna().any()

        valid_rows = group_rows[
            (group_rows['Variation'].notna()) &
            (group_rows['Price'] > 0)
        ]

        valid_count = len(valid_rows)

        if len(group_rows) == 1 and group_rows.iloc[0]['Price'] == 0:
            actions[group_rows.index[0]] = "DELETE"
            continue

        # CASE: single valid variation → collapse (no concatenation of variation name)
        if has_variation and valid_count == 1:
            valid_idx = valid_rows.index[0]

            # ✅ FIX 3: Do NOT concatenate variation name — keep original Item Name only
            df.at[valid_idx, 'Variation'] = ""

            for i in group_rows.index:
                actions[i] = "DELETE"

            actions[valid_idx] = "CONVERTED"
            continue

        # Other cases
        for i, row in group_rows.iterrows():
            if pd.notna(row['Variation']) and row['Price'] == 0:
                actions[i] = "DELETE"
                continue
            if not has_variation and row['Price'] == 0:
                actions[i] = "DELETE"
                continue
            if has_variation and valid_count == 0:
                actions[i] = "DELETE"
                continue

    return actions, df


# ── ROUTES ─────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({"error": "No file uploaded"})

        file_bytes = file.read()
        wb = load_workbook(io.BytesIO(file_bytes))
        protected_rows = get_protected_variation_rows(wb.active, get_ws_columns(wb.active))

        df = pd.read_excel(io.BytesIO(file_bytes))
        df.columns = df.columns.str.strip()

        actions, df_modified = process_logic(df.copy(), protected_rows=protected_rows)

        df_original_raw = df_modified.copy()
        df_original_raw['_status'] = actions

        df_cleaned_raw = df_original_raw[df_original_raw['_status'] != "DELETE"].copy()
        df_items_only = df_cleaned_raw[df_cleaned_raw['Item ID'].isna()].copy().fillna("")
        df_original = df_original_raw.fillna("")
        df_cleaned = df_cleaned_raw.fillna("")

        return jsonify({
            "original": df_original.to_dict(orient='records'),
            "cleaned": df_cleaned.to_dict(orient='records'),
            "items_only": df_items_only.to_dict(orient='records'),
            "deleted_count": actions.count("DELETE"),
            "converted_count": actions.count("CONVERTED")
        })

    except Exception as e:
        return jsonify({"error": str(e)})


@app.route('/download', methods=['POST'])
def download():
    file = request.files['file']
    file_bytes = file.read()

    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = df.columns.str.strip()
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    col_map = get_ws_columns(ws)
    item_name_col = col_map.get('Item Name', 3)
    variation_col = col_map.get('Variation', 4)
    item_type_col = col_map.get('Item Type', 7)
    price_col = col_map.get('Price', 8)
    base_name_col = col_map.get('__base_name', ws.max_column)
    protected_rows = get_protected_variation_rows(ws, col_map)
    actions, df_modified = process_logic(df.copy(), protected_rows=protected_rows)

    num_cols = ws.max_column

    # ── FIX 4: Color fix + reorder (base item above variations, match color) ──
    # We need to do this BEFORE deleting rows, working on original row indices

    # Build a map: base_name → list of (original_row_idx_1based, item_type)
    # We'll find groups by __base_name column (col 13) or by Item Name grouping
    # Since we work on ws directly, let's scan rows grouped by base_name (col 13)

    # Step 1: Read all rows with their colors
    row_info = []  # (excel_row, base_name, item_type, price, bg_color)
    for r in range(2, ws.max_row + 1):
        base_name = ws.cell(r, base_name_col).value
        item_type = ws.cell(r, item_type_col).value
        price = safe_price(ws.cell(r, price_col).value)
        bg = get_cell_bg(ws, r, item_name_col)
        row_info.append({
            'excel_row': r,
            'base_name': str(base_name) if base_name else '',
            'item_type': str(item_type) if item_type else '',
            'price': price,
            'bg': bg
        })

    # Step 2: Group by base_name and fix colors + order
    from collections import defaultdict
    groups = defaultdict(list)
    for info in row_info:
        if info['base_name']:
            groups[info['base_name']].append(info)

    # For groups with mixed colors where the blue row (FF9FC5E8) is "Full" variation
    # and red rows (FFEA9999) are the rest:
    # → Set base item color = blue row color
    # → Reorder: base item first, then variations

    rows_to_reorder = []  # list of (group rows sorted correctly)

    for base_name, grp in groups.items():
        if len(grp) <= 1:
            continue

        base_items = [r for r in grp if r['item_type'] == 'item' and r['price'] == 0]
        variation_items = [r for r in grp if r['item_type'] != 'item']

        if not base_items or not variation_items:
            continue

        colored_variations = [item for item in variation_items if item['bg'] != '00000000']
        representative_variation = sorted(
            colored_variations or variation_items,
            key=lambda item: item['excel_row']
        )[0]
        for r in base_items:
            if representative_variation['bg'] == '00000000':
                continue
            if r['bg'] == representative_variation['bg']:
                continue
            set_row_bg(ws, r['excel_row'], representative_variation['bg'], num_cols)
            r['bg'] = representative_variation['bg']

        # ✅ FIX 4b: Ensure base item (item_type='item') comes FIRST in its group
        # Check if base item is below any variation rows
        base_row = base_items[0]['excel_row']
        # Check if any variation row is ABOVE the base item
        variation_rows_above = [v for v in variation_items if v['excel_row'] < base_row]

        if variation_rows_above:
            rows_to_reorder.append({
                'base': base_items,
                'variations_above': variation_rows_above,
                'variations_below': [v for v in variation_items if v['excel_row'] > base_row],
                'base_row': base_row
            })

    # ── Apply name + variation changes FIRST (before reordering rows) ──
    for i, act in enumerate(actions):
        if act in ["CONVERTED", "KEEP"]:
            new_name = df_modified.iloc[i]['Item Name']
            new_var = df_modified.iloc[i]['Variation']
            if pd.notna(new_name):
                row = i + 2
                ws.cell(row=row, column=item_name_col).value = new_name
                ws.cell(row=row, column=variation_col).value = new_var if pd.notna(new_var) else ""

    # Step 3: Reorder rows where base item is below variations
    # We do this by swapping cell values and colors
    def swap_rows(ws, r1, r2, num_cols):
        """Swap all cell values, fills between two rows."""
        for c in range(1, num_cols + 1):
            c1 = ws.cell(r1, c)
            c2 = ws.cell(r2, c)

            # swap values
            v1, v2 = c1.value, c2.value
            c1.value, c2.value = v2, v1

            # swap fill
            f1_type = c1.fill.fill_type if c1.fill else None
            f1_fg = c1.fill.fgColor.rgb if c1.fill and c1.fill.fgColor and c1.fill.fgColor.type == 'rgb' else None
            f2_type = c2.fill.fill_type if c2.fill else None
            f2_fg = c2.fill.fgColor.rgb if c2.fill and c2.fill.fgColor and c2.fill.fgColor.type == 'rgb' else None

            if f2_fg:
                c1.fill = PatternFill(fill_type='solid', fgColor=f2_fg)
            else:
                c1.fill = PatternFill(fill_type=None)

            if f1_fg:
                c2.fill = PatternFill(fill_type='solid', fgColor=f1_fg)
            else:
                c2.fill = PatternFill(fill_type=None)

    def swap_action_rows(action_list, r1, r2):
        """Keep action indices aligned with worksheet row swaps."""
        idx1 = r1 - 2
        idx2 = r2 - 2
        action_list[idx1], action_list[idx2] = action_list[idx2], action_list[idx1]

    # For each group needing reorder: move base item above the highest variation
    for reorder_group in rows_to_reorder:
        base_row = reorder_group['base_row']
        # Sort variations above base by excel_row
        v_above = sorted(reorder_group['variations_above'], key=lambda x: x['excel_row'])
        top_variation_row = v_above[0]['excel_row']

        # We need to move base_row to just before top_variation_row
        # Do this by successive swaps upward
        current = base_row
        while current > top_variation_row:
            swap_rows(ws, current, current - 1, num_cols)
            swap_action_rows(actions, current, current - 1)
            current -= 1

    # ── ✅ FIX 2: Don't touch orange (FFFFFF00) variation rows ──
    # Delete rows marked DELETE, but skip if Price cell is orange (FFFFFF00)
    rows_to_remove = []
    for i, act in enumerate(actions):
        if act == "DELETE":
            excel_row = i + 2
            if not is_orange_price_row(ws, excel_row, col_map):
                rows_to_remove.append(excel_row)

    for r in sorted(rows_to_remove, reverse=True):
        ws.delete_rows(r)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{os.path.splitext(file.filename)[0]}_cleaned.xlsx"
    return send_file(output, as_attachment=True, download_name=filename)


@app.route('/download-items-only', methods=['POST'])
def download_items_only():
    file = request.files['file']
    file_bytes = file.read()

    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = df.columns.str.strip()

    wb = load_workbook(io.BytesIO(file_bytes))
    protected_rows = get_protected_variation_rows(wb.active, get_ws_columns(wb.active))
    actions, df_modified = process_logic(df.copy(), protected_rows=protected_rows)
    df_modified['_status'] = actions

    df_items = df_modified[
        (df_modified['Item ID'].isna()) &
        (df_modified['_status'] != "DELETE")
    ].copy()

    output_df = pd.DataFrame()
    output_df['Name'] = df_items['Item Name']
    output_df['Item_Online_DisplayName'] = df_items['Online Display Name'].fillna(df_items['Item Name'])
    output_df['Variation_Name'] = df_items['Variation'].fillna("")
    output_df['Price'] = pd.to_numeric(df_items['Price'], errors='coerce').fillna(0)
    output_df['Category'] = df_items['Category']
    output_df['Category_Online_DisplayName'] = df_items['Category']
    output_df['Short_Code'] = ""
    output_df['Short_Code_2'] = ""
    output_df['Description'] = df_items['Description'].fillna("")
    output_df['Attributes'] = df_items['Dietary'].fillna("")
    output_df['Goods_Services'] = "Services"

    output = io.StringIO()
    output_df.to_csv(output, index=False)

    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        as_attachment=True,
        download_name="items_only.csv",
        mimetype='text/csv'
    )


# ── RUN ─────────────────────────────────────────────
if __name__ == '__main__':
    app.run(debug=True)
